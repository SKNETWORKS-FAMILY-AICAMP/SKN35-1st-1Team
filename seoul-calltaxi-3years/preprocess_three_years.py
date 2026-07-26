from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"
REJECTED_DIR = BASE_DIR / "data" / "rejected"

CSV_CHUNK_SIZE = 100_000
EXCEL_BATCH_SIZE = 50_000

RAW_COLUMNS = [
    "접수일시",
    "예정일시",
    "배차일시",
    "승차일시",
    "하차일시",
    "취소일시",
    "출발구",
    "출발동",
    "목적구",
    "목적동",
    "이용목적",
    "요금",
    "승차거리",
    "차량구분",
    "장애유형",
]

COLUMN_ALIASES = {
    "출발지구": "출발구",
    "출발지동": "출발동",
    "목적지구": "목적구",
    "목적지동": "목적동",
    "유형": "장애유형",
}

DATETIME_RENAME = {
    "접수일시": "request_at",
    "예정일시": "scheduled_at",
    "배차일시": "dispatch_at",
    "승차일시": "pickup_at",
    "하차일시": "dropoff_at",
    "취소일시": "cancel_at",
}

OUTPUT_COLUMNS = [
    "source_year",
    "source_row_no",
    "request_at",
    "scheduled_at",
    "dispatch_at",
    "pickup_at",
    "dropoff_at",
    "cancel_at",
    "origin_district_raw",
    "origin_dong_raw",
    "destination_district_raw",
    "destination_dong_raw",
    "purpose_raw",
    "purpose_group",
    "reservation_type",
    "is_reserved",
    "fare",
    "distance_meter",
    "distance_km",
    "vehicle_type",
    "disability_type_raw",
    "disability_type_group",
    "trip_status",
    "origin_is_seoul",
    "destination_is_seoul",
    "route_type",
    "request_date",
    "request_year",
    "request_month",
    "request_hour",
    "request_weekday_num",
    "scheduled_date",
    "scheduled_hour",
    "scheduled_weekday_num",
    "dispatch_wait_min",
    "pickup_wait_from_request_min",
    "pickup_delay_from_schedule_min",
    "dispatch_to_pickup_min",
    "trip_duration_min",
    "time_order_error_flag",
    "extreme_dispatch_wait_flag",
    "extreme_pickup_wait_flag",
    "extreme_dispatch_to_pickup_flag",
    "extreme_trip_duration_flag",
    "completed_missing_fare_flag",
    "completed_zero_distance_flag",
    "extreme_fare_flag",
    "extreme_distance_flag",
    "valid_dispatch_wait_flag",
    "valid_schedule_delay_flag",
    "valid_trip_duration_flag",
    "valid_fare_flag",
    "valid_distance_flag",
]

SEOUL_DISTRICTS = {
    "강남구",
    "강동구",
    "강북구",
    "강서구",
    "관악구",
    "광진구",
    "구로구",
    "금천구",
    "노원구",
    "도봉구",
    "동대문구",
    "동작구",
    "마포구",
    "서대문구",
    "서초구",
    "성동구",
    "성북구",
    "송파구",
    "양천구",
    "영등포구",
    "용산구",
    "은평구",
    "종로구",
    "중구",
    "중랑구",
}

DISABILITY_GROUP_MAP = {
    "뇌병": "뇌병변",
    "일시": "일시적 장애",
    "외국": "외국인",
}


def normalize_column_name(value: object) -> str:
    return str(value).strip().replace("\ufeff", "")


def normalize_raw_frame(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.copy()
    frame.columns = [
        normalize_column_name(column)
        for column in frame.columns
    ]
    frame = frame.rename(columns=COLUMN_ALIASES)

    duplicated = frame.columns[frame.columns.duplicated()].tolist()
    if duplicated:
        raise ValueError(f"중복 컬럼이 있습니다: {duplicated}")

    missing = [
        column
        for column in RAW_COLUMNS
        if column not in frame.columns
    ]
    if missing:
        raise ValueError(
            f"필수 컬럼이 없습니다: {missing}\n"
            f"실제 컬럼: {frame.columns.tolist()}"
        )

    return frame.loc[:, RAW_COLUMNS].copy()


def clean_text(series: pd.Series) -> pd.Series:
    result = series.astype("string").str.strip()
    return result.replace(
        {
            "": pd.NA,
            "nan": pd.NA,
            "NaN": pd.NA,
            "None": pd.NA,
            "NULL": pd.NA,
        }
    )


def clean_number(series: pd.Series) -> pd.Series:
    text = series.astype("string").str.replace(
        ",",
        "",
        regex=False,
    )
    return pd.to_numeric(text, errors="coerce")


def parse_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(
        series,
        errors="coerce",
        format="mixed",
    )


def classify_trip_status(frame: pd.DataFrame) -> pd.Series:
    completed = (
        frame["pickup_at"].notna()
        & frame["dropoff_at"].notna()
    )
    cancelled_after_dispatch = (
        frame["cancel_at"].notna()
        & frame["dispatch_at"].notna()
    )
    cancelled_before_dispatch = frame["cancel_at"].notna()
    missing_pickup = (
        frame["dropoff_at"].notna()
        & frame["pickup_at"].isna()
    )
    missing_dropoff = (
        frame["pickup_at"].notna()
        & frame["dropoff_at"].isna()
    )
    incomplete_after_dispatch = frame["dispatch_at"].notna()

    return pd.Series(
        np.select(
            [
                completed,
                cancelled_after_dispatch,
                cancelled_before_dispatch,
                missing_pickup,
                missing_dropoff,
                incomplete_after_dispatch,
            ],
            [
                "completed",
                "cancelled_after_dispatch",
                "cancelled_before_dispatch",
                "invalid_missing_pickup",
                "invalid_missing_dropoff",
                "incomplete_after_dispatch",
            ],
            default="incomplete_before_dispatch",
        ),
        index=frame.index,
        dtype="string",
    )


def add_route_columns(frame: pd.DataFrame) -> None:
    origin_is_seoul = frame["origin_district_raw"].isin(
        SEOUL_DISTRICTS
    )
    destination_is_seoul = (
        frame["destination_district_raw"].isin(
            SEOUL_DISTRICTS
        )
    )

    frame["origin_is_seoul"] = origin_is_seoul.astype("int8")
    frame["destination_is_seoul"] = (
        destination_is_seoul.astype("int8")
    )
    frame["route_type"] = np.select(
        [
            origin_is_seoul & destination_is_seoul,
            origin_is_seoul & ~destination_is_seoul,
            ~origin_is_seoul & destination_is_seoul,
        ],
        [
            "seoul_to_seoul",
            "seoul_to_outside",
            "outside_to_seoul",
        ],
        default="outside_to_outside",
    )


def preprocess_chunk(
    chunk: pd.DataFrame,
    source_year: int,
    row_offset: int,
) -> pd.DataFrame:
    raw = normalize_raw_frame(chunk)
    frame = pd.DataFrame(index=raw.index)

    frame["source_year"] = source_year
    frame["source_row_no"] = np.arange(
        row_offset + 1,
        row_offset + len(raw) + 1,
        dtype=np.int64,
    )

    for raw_name, output_name in DATETIME_RENAME.items():
        frame[output_name] = parse_datetime(raw[raw_name])

    frame["origin_district_raw"] = clean_text(raw["출발구"])
    frame["origin_dong_raw"] = clean_text(raw["출발동"])
    frame["destination_district_raw"] = clean_text(raw["목적구"])
    frame["destination_dong_raw"] = clean_text(raw["목적동"])

    frame["purpose_raw"] = clean_text(raw["이용목적"])
    reserved = frame["purpose_raw"].str.startswith(
        "예약",
        na=False,
    )
    frame["purpose_group"] = (
        frame["purpose_raw"]
        .str.replace(r"^예약", "", regex=True)
        .replace("", pd.NA)
    )
    frame["reservation_type"] = np.where(
        reserved,
        "reserved",
        "immediate",
    )
    frame["is_reserved"] = reserved.astype("int8")

    frame["fare"] = clean_number(raw["요금"])
    frame["distance_meter"] = clean_number(raw["승차거리"])
    frame["distance_km"] = frame["distance_meter"] / 1000

    frame["vehicle_type"] = clean_text(raw["차량구분"])
    frame["disability_type_raw"] = clean_text(raw["장애유형"])
    frame["disability_type_group"] = (
        frame["disability_type_raw"]
        .replace(DISABILITY_GROUP_MAP)
    )

    frame["trip_status"] = classify_trip_status(frame)
    add_route_columns(frame)

    frame["request_date"] = frame["request_at"].dt.date
    frame["request_year"] = frame["request_at"].dt.year
    frame["request_month"] = frame["request_at"].dt.month
    frame["request_hour"] = frame["request_at"].dt.hour
    frame["request_weekday_num"] = (
        frame["request_at"].dt.weekday
    )

    frame["scheduled_date"] = frame["scheduled_at"].dt.date
    frame["scheduled_hour"] = frame["scheduled_at"].dt.hour
    frame["scheduled_weekday_num"] = (
        frame["scheduled_at"].dt.weekday
    )

    frame["dispatch_wait_min"] = (
        frame["dispatch_at"] - frame["request_at"]
    ).dt.total_seconds() / 60
    frame["pickup_wait_from_request_min"] = (
        frame["pickup_at"] - frame["request_at"]
    ).dt.total_seconds() / 60
    frame["pickup_delay_from_schedule_min"] = (
        frame["pickup_at"] - frame["scheduled_at"]
    ).dt.total_seconds() / 60
    frame["dispatch_to_pickup_min"] = (
        frame["pickup_at"] - frame["dispatch_at"]
    ).dt.total_seconds() / 60
    frame["trip_duration_min"] = (
        frame["dropoff_at"] - frame["pickup_at"]
    ).dt.total_seconds() / 60

    time_order_error = (
        (
            frame["dispatch_at"].notna()
            & (frame["dispatch_at"] < frame["request_at"])
        )
        | (
            frame["pickup_at"].notna()
            & frame["dispatch_at"].notna()
            & (frame["pickup_at"] < frame["dispatch_at"])
        )
        | (
            frame["dropoff_at"].notna()
            & frame["pickup_at"].notna()
            & (frame["dropoff_at"] < frame["pickup_at"])
        )
        | (
            frame["cancel_at"].notna()
            & (frame["cancel_at"] < frame["request_at"])
        )
    )
    frame["time_order_error_flag"] = (
        time_order_error.astype("int8")
    )

    frame["extreme_dispatch_wait_flag"] = (
        (frame["dispatch_wait_min"] < 0)
        | (frame["dispatch_wait_min"] > 720)
    ).fillna(False).astype("int8")
    frame["extreme_pickup_wait_flag"] = (
        (frame["pickup_wait_from_request_min"] < 0)
        | (frame["pickup_wait_from_request_min"] > 720)
    ).fillna(False).astype("int8")
    frame["extreme_dispatch_to_pickup_flag"] = (
        (frame["dispatch_to_pickup_min"] < 0)
        | (frame["dispatch_to_pickup_min"] > 120)
    ).fillna(False).astype("int8")
    frame["extreme_trip_duration_flag"] = (
        (frame["trip_duration_min"] <= 0)
        | (frame["trip_duration_min"] > 360)
    ).fillna(False).astype("int8")

    completed = frame["trip_status"].eq("completed")
    frame["completed_missing_fare_flag"] = (
        completed & frame["fare"].isna()
    ).astype("int8")
    frame["completed_zero_distance_flag"] = (
        completed & frame["distance_meter"].fillna(0).eq(0)
    ).astype("int8")
    frame["extreme_fare_flag"] = (
        (frame["fare"] < 0)
        | (frame["fare"] > 100_000)
    ).fillna(False).astype("int8")
    frame["extreme_distance_flag"] = (
        (frame["distance_meter"] < 0)
        | (frame["distance_meter"] > 300_000)
    ).fillna(False).astype("int8")

    frame["valid_dispatch_wait_flag"] = (
        frame["reservation_type"].eq("immediate")
        & frame["dispatch_wait_min"].between(0, 720)
        & frame["time_order_error_flag"].eq(0)
    ).astype("int8")
    frame["valid_schedule_delay_flag"] = (
        frame["is_reserved"].eq(1)
        & frame["pickup_delay_from_schedule_min"].between(
            -120,
            720,
        )
        & frame["time_order_error_flag"].eq(0)
    ).astype("int8")
    frame["valid_trip_duration_flag"] = (
        completed
        & frame["trip_duration_min"].between(
            0.000001,
            360,
        )
        & frame["time_order_error_flag"].eq(0)
    ).astype("int8")
    frame["valid_fare_flag"] = (
        completed
        & frame["fare"].gt(0)
        & frame["fare"].le(100_000)
    ).astype("int8")
    frame["valid_distance_flag"] = (
        completed
        & frame["distance_meter"].gt(0)
        & frame["distance_meter"].le(300_000)
    ).astype("int8")

    return frame.loc[:, OUTPUT_COLUMNS]


def excel_batches(
    file_path: Path,
    batch_size: int,
) -> Iterator[tuple[str, pd.DataFrame]]:
    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue

            columns = [
                normalize_column_name(value)
                for value in header
            ]
            batch: list[tuple[object, ...]] = []

            for row in rows:
                if not row or not any(
                    value is not None
                    for value in row
                ):
                    continue

                batch.append(row)

                if len(batch) >= batch_size:
                    yield (
                        worksheet.title,
                        pd.DataFrame(batch, columns=columns),
                    )
                    batch = []

            if batch:
                yield (
                    worksheet.title,
                    pd.DataFrame(batch, columns=columns),
                )
    finally:
        workbook.close()


def find_source_file(year: int) -> Path:
    extension = "xlsx" if year == 2023 else "csv"
    candidates = list(
        RAW_DIR.glob(f"*{year}*.{extension}")
    )

    if len(candidates) != 1:
        raise RuntimeError(
            f"{year}년 원본 파일이 {len(candidates)}개입니다: "
            f"{candidates}"
        )

    return candidates[0]


def write_chunk(
    frame: pd.DataFrame,
    file_path: Path,
    first_write: bool,
) -> None:
    frame.to_csv(
        file_path,
        mode="w" if first_write else "a",
        header=first_write,
        index=False,
        encoding="utf-8-sig" if first_write else "utf-8",
        na_rep=r"\N",
        date_format="%Y-%m-%d %H:%M:%S.%f",
        lineterminator="\n",
    )


def process_year(year: int) -> None:
    source_file = find_source_file(year)
    output_file = (
        PROCESSED_DIR
        / f"taxi_trip_cleaned_{year}.csv"
    )
    rejected_file = (
        REJECTED_DIR
        / f"taxi_trip_rejected_{year}.csv"
    )

    output_file.unlink(missing_ok=True)
    rejected_file.unlink(missing_ok=True)

    row_offset = 0
    rejected_count = 0
    first_output = True
    first_rejected = True

    if year == 2023:
        reader = (
            (sheet_name, chunk)
            for sheet_name, chunk in excel_batches(
                source_file,
                EXCEL_BATCH_SIZE,
            )
        )
    else:
        reader = (
            ("CSV", chunk)
            for chunk in pd.read_csv(
                source_file,
                encoding="utf-8-sig",
                chunksize=CSV_CHUNK_SIZE,
                low_memory=False,
            )
        )

    for chunk_number, (source_part, chunk) in enumerate(
        reader,
        start=1,
    ):
        cleaned = preprocess_chunk(
            chunk=chunk,
            source_year=year,
            row_offset=row_offset,
        )

        rejected_mask = (
            cleaned["trip_status"].str.startswith(
                "invalid",
                na=False,
            )
            | cleaned["time_order_error_flag"].eq(1)
        )
        rejected = cleaned.loc[rejected_mask]

        write_chunk(cleaned, output_file, first_output)
        first_output = False

        if not rejected.empty:
            write_chunk(
                rejected,
                rejected_file,
                first_rejected,
            )
            first_rejected = False
            rejected_count += len(rejected)

        row_offset += len(chunk)
        print(
            f"{year}년 {source_part} "
            f"{chunk_number}번째 청크 완료 "
            f"| 누적 {row_offset:,}행 "
            f"| 검토 대상 {rejected_count:,}행"
        )

    print(
        f"{year}년 완료 | 전체 {row_offset:,}행 "
        f"| 검토 대상 {rejected_count:,}행"
    )
    print(f"전처리 파일: {output_file}")
    if rejected_count:
        print(f"검토 파일: {rejected_file}")


def main() -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    REJECTED_DIR.mkdir(parents=True, exist_ok=True)

    for year in (2023, 2024, 2025):
        process_year(year)

    print("3개년 전처리가 모두 완료되었습니다.")


if __name__ == "__main__":
    main()
