from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
RAW_DIR = BASE_DIR / "data" / "raw"

EXPECTED_COLUMNS = [
    "접수일시",
    "예정일시",
    "배차일시",
    "승차일시",
    "하차일시",
    "취소일시",
    "출발지구",
    "출발지동",
    "목적지구",
    "목적지동",
    "이용목적",
    "요금",
    "승차거리",
    "차량구분",
    "장애유형",
]


def normalize_columns(columns):
    return [
        str(column).strip().replace("\ufeff", "")
        for column in columns
    ]


def find_one(pattern):
    files = list(RAW_DIR.glob(pattern))

    if len(files) != 1:
        raise RuntimeError(
            f"'{pattern}'에 해당하는 파일이 "
            f"{len(files)}개입니다: {files}"
        )

    return files[0]


def inspect_excel():
    file_path = find_one("*2023*.xlsx")

    print("=" * 70)
    print("2023년 Excel:", file_path.name)

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    try:
        print("시트 목록:", workbook.sheetnames)

        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)

            if header is None:
                print(f"- {worksheet.title}: 빈 시트")
                continue

            columns = normalize_columns(header)

            print(f"\n- 시트: {worksheet.title}")
            print("  예상 행 수:", f"{max(worksheet.max_row - 1, 0):,}")
            print("  열 수:", len(columns))
            print("  컬럼:", columns)
            print("  기준 컬럼 일치:", columns == EXPECTED_COLUMNS)

    finally:
        workbook.close()


def inspect_csv(year):
    file_path = find_one(f"*{year}*.csv")

    print("=" * 70)
    print(f"{year}년 CSV:", file_path.name)

    sample = pd.read_csv(
        file_path,
        encoding="utf-8-sig",
        nrows=5,
        low_memory=False,
    )

    columns = normalize_columns(sample.columns)

    print("열 수:", len(columns))
    print("컬럼:", columns)
    print("기준 컬럼 일치:", columns == EXPECTED_COLUMNS)
    print("\n처음 2행:")
    print(sample.head(2).to_string(index=False))


def main():
    print("원본 폴더:", RAW_DIR)
    print("원본 폴더 존재:", RAW_DIR.exists())

    inspect_excel()
    inspect_csv(2024)
    inspect_csv(2025)

    print("=" * 70)
    print("원본 구조 검사 완료")


if __name__ == "__main__":
    main()